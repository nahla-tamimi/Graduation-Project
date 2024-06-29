from flask import Flask, request, jsonify
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
from openpyxl import Workbook

app = Flask(__name__)

# Connect to the MySQL database
connection = mysql.connector.connect(

    host="localhost",
    user="optifisc_hireadmin",
    passwd="A$C.cC9W}}i9",
    database="optifisc_hire"

)

if connection.is_connected():
    print("MySQL Connected...")
else:
    print("Connection Failed.")

cursor = connection.cursor()


@app.route("/")
def index():
    return "Welcome to the Flask MySQL Admin API!"


@app.route("/test")
def home():
    return "Home"


# Endpoint to get information of a specific admin by id
@app.route("/get-admins/<id>")
def get_admins(id):
    # Execute the query to retrieve admin information from the database
    cursor.execute("SELECT id, name, email FROM admins WHERE id=%s", (id,))
    admins_data = cursor.fetchone()

    if admins_data:
        # Convert the data to JSON format and return it
        admins_json = {
            "id": admins_data[0],
            "name": admins_data[1],
            "email": admins_data[2]
        }
        return jsonify(admins_json), 200
    else:
        return "Admin not found", 404


@app.route("/get-candidate/<id>")
def get_candidate(id):
    # Execute the query to retrieve candidate information from the database
    cursor.execute("SELECT id, first_name, last_name, bio FROM candidates WHERE id=%s", (id,))
    candidate_data = cursor.fetchone()

    if candidate_data:
        # Convert the data to JSON format and return it
        candidate_json = {
            "id": candidate_data[0],
            "first_name": candidate_data[1],
            "last_name": candidate_data[2],
            "bio": candidate_data[3]
        }
        return jsonify(candidate_json), 200
    else:
        return "Candidate not found", 404


@app.route("/get-candidates-with-skills")
def get_candidates_with_skills():
    try:
        # Execute the SQL query
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all the results
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Create a list to store the results
            results = []
            # Iterate over the retrieved data and format it
            for candidate_data in candidates_data:
                result = {
                    "first_name": candidate_data[0],
                    "last_name": candidate_data[1],
                    "email": candidate_data[2],
                    "address": candidate_data[3],
                    "bio": candidate_data[4],
                    "skills": candidate_data[5].split(','),  # Convert skills string to list
                    "predicted_job_category": predict_job_category(candidate_data[4] + " " + candidate_data[5])

                }
                results.append(result)

            # Return the formatted data as JSON
            return jsonify(results), 200
        else:
            return "No candidates with applied jobs found", 404
    except mysql.connector.Error as error:
        return f"Error fetching data from the database: {error}", 500


from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import LabelEncoder
import nltk
import os
import pandas as pd
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import string
from flask import Flask, request, jsonify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import pickle
import nltk

from joblib import load
# nltk.download(download_dir="nltk_data")
# # import nltk
# nltk.download()
# Load the model from the file
# Load the model from the file
loaded_model = load("model_1.4.2.joblib")

# Load the word vectorizer from the file
word_vectorizer = load("word_vectorizer_1.4.2.joblib")

# Load the encoder from the file
Encoder = load("Encoder_1.4.2.joblib")


# Function to predict job category
def predict_job_category(resume_text):
    resume_vectorized = word_vectorizer.transform([resume_text])
    predicted_category_index = loaded_model.predict(resume_vectorized)[0]
    predicted_category = Encoder.inverse_transform([predicted_category_index])[0]
    return predicted_category


def preprocess_text(text):
    tokens = word_tokenize(text.lower())
    stop_words = set(stopwords.words('english'))
    punctuation = set(string.punctuation)
    filtered_tokens = [word for word in tokens if word not in stop_words and word not in punctuation]
    preprocessed_text = ' '.join(filtered_tokens)
    return preprocessed_text


def find_score(jobdes, resume, customKeywords):
    resume = preprocess_text(resume)
    custom_keywords = ' '.join(customKeywords)
    jobdes = jobdes + ' ' + custom_keywords
    text = [resume, jobdes]
    cv = CountVectorizer()
    count_matrix = cv.fit_transform(text)
    matchpercent = cosine_similarity(count_matrix)[0][1] * 100
    matchpercent = round(matchpercent, 2)
    return matchpercent


def find_best_matching_category(job_description, resume, job_categories, custom_keywords):
    resume = preprocess_text(resume)
    custom_keywords = ' '.join(custom_keywords)
    job_description = job_description + ' ' + custom_keywords

    max_score = 0
    best_category = ""

    for category, keywords in job_categories.items():
        job_keywords = ' '.join(keywords)
        text = [resume, job_keywords]
        cv = CountVectorizer()
        count_matrix = cv.fit_transform(text)
        matchpercent = cosine_similarity(count_matrix)[0][1] * 100
        matchpercent = round(matchpercent, 2)

        if matchpercent > max_score:
            max_score = matchpercent
            best_category = category

    return best_category


def process_resumes(job_categories_files, candidates_df, job_categories):
    results = []
    for index, row in candidates_df.iterrows():
        resume_text = row['bio']
        job_description_scores = {}
        for category, job_description_file in job_categories_files.items():
            job_description = open(job_description_file, 'r', encoding='utf-8').read()
            score = find_score(job_description, resume_text, job_categories[category])
            job_description_scores[category] = score

        best_category = max(job_description_scores, key=job_description_scores.get)
        data_science_score = job_description_scores.get('Data_Science', 0)
        database_score = job_description_scores.get('Database', 0)
        web_design_score = job_description_scores.get('Web_Design', 0)
        network_security_score = job_description_scores.get('Network_Security_Engineer', 0)

        results.append({
            'Index': index,
            'Resume': resume_text,
            'Result': best_category,
            'Data Science Score': data_science_score,
            'Database Score': database_score,
            'Web Design Score': web_design_score,
            'Network Security Score': network_security_score
        })

    return pd.DataFrame(results)


job_categories_files = {
    "Data Science": r'Data_Science.txt',
    "Database": r'Database.txt',
    "Web Designing": r'Web_Design.txt',
    "Network Security Engineer": r'Network_Security_Engineer.txt'
}

job_categories = {
    "Data Science": ["python", "machine learning", "big data", "deep learning", "ai", "artificial intelligence",
                     "data analysis", "statistical modeling", "visualization", "regression", "cluster analysis",
                     "sentiment analysis", "natural language processing"],
    "Database": ["mysql", "sql", "oracle", "relational database management system", "database design", "normalization",
                 "query", "replication", "optimization", "integrity", "dbms", "data access"],
    "Web Designing": ["web design", "design", "ui/ux", "html", "prototyping", "css", "javascript", "node.js",
                      "react.js", "laravel", "web2py", "asp.net", "angular js", "html5", "css3", "sass", "bootstrap",
                      "jquery", "javascript", "visual studio", "photoshop"],
    "Network Security Engineer": ["computer engineering", "cybersecurity", "secure", "networks", "firewall",
                                  "detection", "prevention", "protocols", "encryption", "cissp", "cisco", "router",
                                  "server", "lan", "wan", "access control", "authentication", "host"]
}


@app.route("/process-resumes")
def process_resumes_endpoint():
    try:
        # Execute the SQL query to retrieve candidate data
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all the results
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Create a DataFrame to store the results
            candidates_df = pd.DataFrame(candidates_data,
                                         columns=['first_name', 'last_name', 'email', 'address', 'bio', 'skills'])

            # Process the resumes
            results_df = process_resumes(job_categories_files, candidates_df, job_categories)

            # Convert DataFrame to JSON with each candidate on a new line
            results_json = results_df.to_json(orient="records", lines=True)

            return results_json, 200
        else:
            return "No candidates with applied jobs found", 404
    except mysql.connector.Error as error:
        return f"Error fetching data from the database: {error}", 500


# Function to highlight top resumes for each job category
def highlight_top_resumes(df):
    # Dictionary to map job categories to colors
    job_colors = {
        'Data Science': "FF0000",  # Red
        'Database': "FFFF00",  # Yellow
        'Web Design': "008000",  # Green
        'Network Security': "0000FF"  # Blue
    }

    # Create a new workbook
    workbook = Workbook()
    # Get the active sheet
    sheet = workbook.active

    # Function to highlight top resumes in a specific column with a given fill color
    def highlight_top(column_name, fill_color):
        top_resumes = df.nlargest(5, column_name)
        for idx, row in top_resumes.iterrows():
            row_number = idx + 2  # Excel rows start from 1, but DataFrame indices start from 0, so we add 2 to match Excel rows
            # تحديد الخلية وتلوينها باللون المحدد
            cell = sheet.cell(row=row_number, column=df.columns.get_loc(column_name) + 1)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # تمرير اسماء الخانات 
    for job_title, color in job_colors.items():
        highlight_top(job_title + ' Score', color)

    new_file_path = r'datahighlighted_resumes.xlsx'
    workbook.save(new_file_path)
    return new_file_path  # Return the path of the newly created Excel file


# Define a new endpoint for processing resumes
@app.route("/process-resumes-highlight")
def process_resumes_new_endpoint():
    try:
        # Execute the SQL query to retrieve candidate data
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all the results
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Create a DataFrame to store the results
            candidates_df = pd.DataFrame(candidates_data,
                                         columns=['first_name', 'last_name', 'email', 'address', 'bio', 'skills'])
            # Process the resumes
            results_df = process_resumes(job_categories_files, candidates_df, job_categories)
            # Convert DataFrame to JSON
            results_json = results_df.to_json(orient="records")

            # Call the highlight_top_resumes function after processing the resumes
            highlighted_excel_file = highlight_top_resumes(results_df)
            print("Highlighted resumes saved to:", highlighted_excel_file)

            return results_json, 200
        else:
            return "No candidates with applied jobs found", 404
    except mysql.connector.Error as error:
        return f"Error fetching data from the database: {error}", 500


# Function to process resumes, calculate match scores, and find top matching candidates for each job
def process_resumes_and_find_top_candidates(job_categories_files, candidates_df, job_categories):
    results = {}

    # Iterate over each job category
    for category, job_description_file in job_categories_files.items():
        job_description = open(job_description_file, 'r', encoding='utf-8').read()
        results[category] = []

        # Iterate over each candidate
        for index, row in candidates_df.iterrows():
            resume_text = preprocess_text(row['bio'])
            candidate_skills = row['skills'].split(',')

            # Calculate the match score between job description and candidate
            match_score = float(find_score(job_description, resume_text, candidate_skills))

            # Append candidate's information and match score to the results list
            results[category].append({
                'name': row['first_name'] + ' ' + row['last_name'],
                'email': row['email'],
                'address': row['address'],
                'bio': resume_text,
                'skills': candidate_skills,
                'category': category,
                'match_score': match_score  # Add match_score here
            })

        # Sort candidates by match score in descending order
        results[category] = sorted(results[category], key=lambda x: x['match_score'], reverse=True)[:5]

    return results


import json


# Define a new endpoint to process resumes and find top candidates for each job
@app.route("/process-resumes-and-find-top-candidates")
def process_resumes_and_find_top_candidates_endpoint():
    try:
        # Execute the SQL query to retrieve candidate data
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all the results
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Create a DataFrame to store the results
            candidates_df = pd.DataFrame(candidates_data,
                                         columns=['first_name', 'last_name', 'email', 'address', 'bio', 'skills'])

            # Process the resumes and find top candidates for each job
            results = process_resumes_and_find_top_candidates(job_categories_files, candidates_df, job_categories)

            # Convert results to JSON format
            results_json = json.dumps(results, indent=4)

            # Return the JSON data
            return results_json, 200
        else:
            return "No candidates with applied jobs found", 404
    except mysql.connector.Error as error:
        return f"Error fetching data from the database: {error}", 500


import json


# Define a new endpoint to process resumes and find top candidates for each job
@app.route("/process-resumes-name-and-score")
def process_resumes_name_and_score_endpoint():
    try:
        # Execute the SQL query to retrieve candidate data
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all the results
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Create a DataFrame to store the results
            candidates_df = pd.DataFrame(candidates_data,
                                         columns=['first_name', 'last_name', 'email', 'address', 'bio', 'skills'])

            # Process the resumes and find top candidates for each job
            results = process_resumes_and_find_top_candidates(job_categories_files, candidates_df, job_categories)

            # Prepare a dictionary to store results for each job category
            results_dict = {}

            # Iterate over each job category
            for category, candidates in results.items():
                # Prepare a list to store candidate names and match scores
                candidates_list = []
                # Iterate over the top candidates for the current job category
                for candidate in candidates:
                    # Append candidate name and match score to the list
                    candidates_list.append({
                        'name': candidate['name'],
                        'match_score': candidate['match_score']
                    })
                # Store the list in the results dictionary
                results_dict[category] = candidates_list

            # Convert results to JSON format
            results_json = json.dumps(results_dict, indent=4)

            # Return the JSON data
            return results_json, 200
        else:
            return "No candidates with applied jobs found", 404
    except mysql.connector.Error as error:
        return f"Error fetching data from the database: {error}", 500


############################################################################################################################


@app.route("/process-resumes-calculate-match-score-store-db")
def process_resumes_calculate_match_score_store_db_endpoint():
    try:
        # Execute the SQL query to retrieve candidate data
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all the results
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Create a DataFrame to store the results
            candidates_df = pd.DataFrame(candidates_data,
                                         columns=['first_name', 'last_name', 'email', 'address', 'bio', 'skills'])

            # Process the resumes and find top candidates for each job
            results = process_resumes_and_find_top_candidates(job_categories_files, candidates_df, job_categories)

            # Iterate over the results and store them in the database
            for category, candidates_info in results.items():
                for candidate_info in candidates_info:
                    name = candidate_info['name']
                    email = candidate_info['email']
                    address = candidate_info['address']
                    bio = candidate_info['bio']
                    skills = ','.join(candidate_info['skills'])
                    category_name = category  # Use the category name as the category
                    match_score = float(candidate_info['match_score'])

                    # Execute the SQL query to insert the results into the database
                    sql = "INSERT INTO candidates_scores (name, email, address, bio, skills, category, match_score) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                    val = (name, email, address, bio, skills, category_name, match_score)
                    cursor.execute(sql, val)

            # Commit the changes
            connection.commit()

            return "Results inserted into the database successfully", 200
        else:
            return "No candidates with applied jobs found", 404
    except mysql.connector.Error as error:
        return f"Error fetching or inserting data into the database: {error}", 500


# Define the endpoint to update predicted job categories and candidate data
@app.route("/update-candidates-scores-and-process-resumes")
def update_candidates_scores_and_process_resumes():
    try:
        # Empty the candidates_scores table
        cursor.execute("TRUNCATE TABLE candidates_scores")

        # Commit the changes
        connection.commit()

        # Call process_resumes_calculate_match_score_store_db_endpoint to process resumes and store data
        response, status_code = process_resumes_calculate_match_score_store_db_endpoint()

        # Check if the operation was successful
        if status_code == 200:
            return "Candidates scores updated and resumes processed successfully", 200
        else:
            return response, status_code

    except mysql.connector.Error as error:
        return f"Error updating candidates scores and processing resumes: {error}", 500


############################################################################################################################


# Define the endpoint to predict job category and store candidate data
@app.route("/predict-job-category-and-store-candidate-data")
def predict_job_category_and_store_candidate_data_endpoint():
    try:
        # Execute the SQL query to retrieve candidate data
        cursor.execute("""
            SELECT 
                u.first_name,
                u.last_name,
                u.email,
                c.address,
                SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                GROUP_CONCAT(st.name) AS skills
            FROM 
                users u
            JOIN 
                candidates c ON u.id = c.user_id
            JOIN 
                candidate_skill cs ON c.id = cs.candidate_id
            JOIN 
                skill_translations st ON cs.skill_id = st.skill_id
            WHERE 
                c.id IN (SELECT candidate_id FROM applied_jobs)
            GROUP BY
                u.id, c.address, c.bio
        """)

        # Fetch all candidates
        candidates_data = cursor.fetchall()

        # Check if any data is retrieved
        if candidates_data:
            # Get column names
            column_names = [i[0] for i in cursor.description]

            # Iterate over each candidate
            for candidate_data in candidates_data:
                candidate_dict = dict(zip(column_names, candidate_data))

                # Extract candidate information
                first_name = candidate_dict['first_name']
                last_name = candidate_dict['last_name']
                email = candidate_dict['email']
                address = candidate_dict['address']
                bio = candidate_dict['bio']
                skills = candidate_dict['skills']

                # Predict job category
                resume_text = bio + " " + skills
                predicted_category = predict_job_category(resume_text)

                # Retrieve job description based on predicted category
                job_description_file = job_categories_files.get(predicted_category)

                # Read job description from file
                job_description = open(job_description_file, 'r', encoding='utf-8').read()

                # Calculate match score
                match_score = find_score(job_description, resume_text, skills)

                # Print candidate information and match score
                print("Candidate Information:")
                print("First Name:", first_name)
                print("Last Name:", last_name)
                print("Email:", email)
                print("Address:", address)
                print("Bio:", bio)
                print("Skills:", skills)
                print("Predicted Category:", str(predicted_category))
                print("Match Score:", float(match_score))

                # Execute the SQL query to insert the results into the database
                sql = "INSERT INTO predicts_scores (first_name, last_name, email, address, bio, skills, predicted_category, match_score) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                cursor.execute(sql,
                               (first_name, last_name, email, address, bio, skills, predicted_category,
                                float(match_score)))

            # Commit the changes
            connection.commit()

            return "Results inserted into the database successfully", 200
        else:
            return "No candidates with applied jobs found", 404

    except mysql.connector.Error as error:
        return f"Error fetching or inserting data into the database: {error}", 500


# Define the endpoint to update predicted job categories and store candidate data
@app.route("/update-predicted-job-categories-and-candidate-data")
def update_predicted_job_categories_and_candidate_data():
    try:
        # Empty the predicts_scores table
        cursor.execute("TRUNCATE TABLE predicts_scores")

        # Commit the changes
        connection.commit()

        # Call predict_job_category_and_store_candidate_data_endpoint to populate predicts_scores
        response, status_code = predict_job_category_and_store_candidate_data_endpoint()

        # Check if the operation was successful
        if status_code == 200:
            return "Predicted job categories and candidate data updated successfully", 200
        else:
            return response, status_code

    except mysql.connector.Error as error:
        return f"Error updating predicted job categories and candidate data: {error}", 500


############################################################################################################################


# Define the endpoint to retrieve top candidates for each job category
@app.route("/top-candidates")
def top_candidates_endpoint():
    try:
        # Initialize a dictionary to store results
        top_candidates = {}

        # Iterate over each job category
        for category, job_description_file in job_categories_files.items():
            # Read job description from file
            job_description = open(job_description_file, 'r', encoding='utf-8').read()

            # Initialize a list to store candidates for this category
            category_candidates = []

            # Fetch candidates for this category from the database
            cursor.execute("""
                SELECT 
                    u.first_name,
                    u.last_name,
                    u.email,
                    c.address,
                    SUBSTRING(c.bio, 5, LENGTH(c.bio) - 8) AS bio,
                    GROUP_CONCAT(st.name) AS skills
                FROM 
                    users u
                JOIN 
                    candidates c ON u.id = c.user_id
                JOIN 
                    candidate_skill cs ON c.id = cs.candidate_id
                JOIN 
                    skill_translations st ON cs.skill_id = st.skill_id
                WHERE 
                    c.id IN (SELECT candidate_id FROM applied_jobs)
                GROUP BY
                    u.id, c.address, c.bio
            """)

            # Fetch all candidates for this category
            candidates_data = cursor.fetchall()

            # Check if any candidates are found
            if candidates_data:
                # Get column names
                column_names = [i[0] for i in cursor.description]

                # Iterate over each candidate
                for candidate_data in candidates_data:
                    # Create a dictionary for the current candidate
                    candidate_dict = dict(zip(column_names, candidate_data))

                    # Extract candidate information
                    first_name = candidate_dict['first_name']
                    last_name = candidate_dict['last_name']
                    email = candidate_dict['email']
                    address = candidate_dict['address']
                    bio = candidate_dict['bio']
                    skills = candidate_dict['skills']

                    # Calculate match score
                    match_score =float(find_score(job_description, bio + " " + skills, skills))

                    # Append candidate's information and match score to the results list
                    category_candidates.append({
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'address': address,
                        'bio': bio,
                        'skills': skills,
                        'category': category,
                        'match_score': match_score
                    })

                # Sort candidates by match score in descending order
                category_candidates = sorted(category_candidates, key=lambda x: x['match_score'], reverse=True)[:5]

                # Add the top candidates for this category to the results dictionary
                top_candidates[category] = category_candidates

                # Insert candidate data into the predict_score table
                for candidate in category_candidates:
                    cursor.execute("""
                        INSERT INTO predict_score (first_name, last_name, email, address, bio, skills, category, match_score)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        candidate['first_name'],
                        candidate['last_name'],
                        candidate['email'],
                        candidate['address'],
                        candidate['bio'],
                        candidate['skills'],
                        candidate['category'],
                        float(candidate['match_score'])
                    ))

        # Commit changes to the database
        connection.commit()
        return "Top candidates stored in the predict_score table successfully"


    except mysql.connector.Error as error:
        return f"Error storing data in the predict_score table: {error}"


# Define the endpoint to update predicted job categories and candidate data
@app.route("/update-top-candidates-data")
def update_top_candidates_data():
    try:
        # Clear the predict_score table
        cursor.execute("TRUNCATE TABLE predict_score")

        # Commit the changes
        connection.commit()

        # Call the top_candidates_endpoint to update the predict_score table
        response = top_candidates_endpoint()

        # Return the response from the top_candidates_endpoint
        return jsonify(response), 200 

    except mysql.connector.Error as error:
        return f"Error updating data in the database: {error}", 500


if __name__ == "__main__":
    app.run(debug=True)
